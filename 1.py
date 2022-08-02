#!/usr/bin/python3
# encoding:utf-8
import time as t
from selenium import webdriver
from openpyxl import load_workbook
import openpyxl
from time import sleep


# 已修改
import pywinauto
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import requests
def readexcel(filename):
    wb = load_workbook(filename)
    # 获取指定表名
    sheet = wb['Sheet1']
    # 获取当前活动表
    # sheet=wb.active
    # 取表头
    data = sheet['A1':'M1']
    list = [] # key值
    for cell in data:
        for key in cell:
            list.append(key.value)
    # 按行读取
    cell_rows=sheet.iter_rows(min_row=2,max_row=sheet.max_row,min_col=1,max_col=sheet.max_column,values_only=True)
    data = []
    # 读取value,与title合并
    for key in cell_rows:
        list1=[]#每行的value
        for cell in key:
            list1.append(cell)
        data.append(dict(zip(list,list1)))
    return data
wa = openpyxl.load_workbook("D:/迅雷下载/数据.xlsx")
sht = wa['Sheet1']
A2 = sht['A1']
a2 = str(A2.value)
print(a2)
data = readexcel("D:/迅雷下载/"+a2)
print(data)
da = data
common_key1 = [item['出货单号'] for item in da]
# print(common_key1)
common_key2 = [item['出库通知书编号'] for item in da]
common_key3 = [item['出货数量'] for item in da]
new_common_key3=[str(i) for i in common_key3]
# print(type(new_common_key3))
common_key4 = [item['备注'] for item in da]
common_key5 = [item['出货时间'] for item in da]
common_key6 = [item['预计到货时间'] for item in da]
common_key7 = [item['电子单据模板'] for item in da]
common_key8 = [item['制单人'] for item in da]
common_key9 = [item['附件'] for item in da]
common_key10 = [item['登录账号'] for item in da]
common_key11 = [item['登录密码'] for item in da]
common_key12 = [item['附件路径'] for item in da]
common_key13 = [item['运行次数'] for item in da]
# print(common_key13)
common = [common_key1, common_key2,new_common_key3,common_key4,common_key5,common_key6,common_key7,common_key8,common_key9,common_key10,common_key11,common_key12,common_key13]
# print(new_common_key3[0][0])
# print(common_key13[12][0])
n=common[12][1]
print(n)
i=0
i=int(i)
while i <= 1:
    driver = webdriver.Chrome(executable_path="chromedriver.exe")
    driver.maximize_window()
    # driver.get("https://www.gxjtyl.com/s/#/login") # 正式服
    driver.get("https://demo.gxjtyl.com/s/#/login") # 测试服
    driver.find_element_by_name("userPhone").send_keys(18778167107)
    driver.implicitly_wait(10)
    driver.find_element_by_name("password").send_keys("123456")
    driver.implicitly_wait(10)
    driver.find_element_by_xpath("/html/body/div/div/section/div/div[2]/form/button").click()
    driver.implicitly_wait(10)
    # # 选择企业
    driver.find_element_by_css_selector(".el-col:nth-child(4) .el-button").click()
    driver.implicitly_wait(10)
    driver.find_element_by_css_selector(".menu-wrapper:nth-child(9) .el-submenu__title").click()
    driver.implicitly_wait(10)
    driver.find_element_by_css_selector(".menu-wrapper:nth-child(9) .menu-wrapper:nth-child(4) .el-menu-item").click()
    # 到货签收单号
    driver.implicitly_wait(10)
    # driver.find_element_by_xpath("//*[@id='app']/div/div[2]/section/div/div[1]/div[2]/input").send_keys("DHD-20211110-12206")
    # sleep(1)
    # driver.find_element_by_xpath("//*[@id='app']/div/div[2]/section/div/div[1]/button[1]").click()
    driver.implicitly_wait(10)
    driver.find_element_by_xpath("//*[@id='app']/div/div[2]/section/div/div[3]/div[3]/table/tbody/tr["+i+"]/td[7]/div/button").click()
    # sleep(1)
    # driver.find_element_by_xpath("/html/body/div[2]/div/div[2]/form/div[4]/div[1]/div/div/div/div[1]/ul/li/img").click()
    # driver.find_element_by_xpath("/html/body/div[2]/div/div[2]/form/div[4]/div[1]/div/div/div/div[1]/ul/li/span")
    # r = requests.get("https://demo.gxjtyl.com/s/#/stock/arrival")
    # print(r.)
    # driver.find_element_by_css_selector(".el-icon-zoom-in").click()
    # driver.find_element_by_class_name("el-upload-list__item-preview").click()
    # driver.find_element_by_xpath("/html/body/div[2]/div/div[2]/form/div[4]/div[1]/div/div/div/div[1]/ul/li/span").click()
    # sleep(2)
    # driver.find_element_by_xpath("/html/body/div[2]/div/div[2]/form/div[4]/div[1]/div/div/div/div[1]/ul/li/span/span[1]").click()
    sleep(2)
    action=driver.find_element_by_xpath("/html/body/div[2]/div/div[2]/form/div[4]/div[1]/div/div/div/div[1]/ul/li/img").get_attribute("src")

    print(action)
    sleep(1)
    r=requests.get(action)
    # ActionChains(driver).context_click(action).move_to_element(action).perform()
    # sleep(1)
    path = 'D:/迅雷下载/xiaz/1.jpg'
    print('正在下载' + action)
    with open(path, 'wb') as f:
        f.write(r.content)
        # f.close()
        print("下载成功")
    sleep(5)
    # # 点击键盘向下箭头
    # driver.find_element_by_xpath().send_keys(Keys.ENTER)
    # sleep(1)
    # driver.send_keys(Keys.DOWN)

    # action.click(Keys,'v')
    # # action.perform()  # 执行保存
    # print(driver.title)
    # t.sleep(5)
    # driver.close()
    i+=1
    if i==1:
        break
